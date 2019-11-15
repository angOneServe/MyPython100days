from openpyxl import Workbook
#创建空工作簿
wb=Workbook()
wb.save("test.xlsx")
#获取工作表 修改其title
ws=wb.active
print(ws.title)
ws.title="actve获取的ws已被修改"
ws.sheet_properties.tabColor="1072BA"
#创建工作表
ws1=wb.create_chartsheet("mysheet1")
ws2=wb.create_chartsheet("mysheet2")
#获取工作表

print(wb["actve获取的ws已被修改"].title)
ws3=wb.copy_worksheet(wb["actve获取的ws已被修改"])

#访问单元格
ws3["a1"]="姓名"
ws3["a2"]="年龄"
print(ws3["a1"].value)

for row in ws3["c1":"g7"]:
    for cell in row:
        print(cell.value)
print("iter_rows")
for col in ws3.iter_rows(min_row=1,max_row=2,min_col=1,max_col=1):
    for v in col:
        print(v.value)
wb.save("test.xlsx")

#获得所有行或者列
for col in ws3.columns:
    print("列")
    for v in col:
        print(v.value)

for r in ws3.rows:
    print("行")
    for v in r:
        print(v.value)
#返回所有值
for v in ws3.values:
    print(v)
for col in ws3.iter_rows(min_row=1,max_row=2,min_col=1,max_col=1,values_only=True):
    for v in col:
        print(v)


#数据存储
for col in ws3.columns:
    print("列")
    for v in col:
        v.value=1
wb.save("test.xlsx")
